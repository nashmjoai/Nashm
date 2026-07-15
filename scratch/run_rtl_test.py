import os
import sys
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from pptx import Presentation
from docx import Document
from openpyxl import Workbook, load_workbook

# Force stdout to use utf-8 representation or print safe ascii
def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', errors='backslashreplace').decode('ascii'))

def run_export(kind, input_json, output_file):
    print(f"\n================ Running export for {kind} ================")
    cmd = [
        "python",
        "packages/api/src/artifacts/export.py",
        "--input", input_json,
        "--output", output_file,
        "--format", "office"
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    print("STDOUT:")
    safe_print(res.stdout)
    if res.stderr:
        print("STDERR:")
        safe_print(res.stderr)
    return res.returncode == 0

def test_slides():
    input_json = "scratch/payload_slides.json"
    output_file = "scratch/test_output.pptx"
    if not run_export("slides", input_json, output_file):
        print("FAIL: Slides export crashed")
        return False
    
    print("\n--- PPTX Verification ---")
    with zipfile.ZipFile(output_file, 'r') as z:
        slide_xml = z.read("ppt/slides/slide1.xml").decode('utf-8')
        print("Does slide1.xml contain rtl=\"1\"?")
        if 'rtl="1"' in slide_xml:
            print("PASS: found rtl=\"1\" in slide xml")
            idx = slide_xml.find('rtl="1"')
            print("XML Snippet around rtl=\"1\":")
            safe_print(slide_xml[max(0, idx-80):min(len(slide_xml), idx+120)])
        else:
            print("FAIL: rtl=\"1\" NOT found in slide xml")
            
def test_document():
    input_json = "scratch/payload_document.json"
    output_file = "scratch/test_output.docx"
    if not run_export("document", input_json, output_file):
        print("FAIL: Document export crashed")
        return False
        
    print("\n--- DOCX Verification ---")
    found_bidi = False
    found_rtl = False
    
    with zipfile.ZipFile(output_file, 'r') as z:
        doc_xml = z.read("word/document.xml").decode('utf-8')
        if '<w:bidi/>' in doc_xml or '<w:bidi' in doc_xml:
            found_bidi = True
            idx = doc_xml.find('<w:bidi')
            print("Found <w:bidi/>. XML Snippet:")
            safe_print(doc_xml[max(0, idx-60):min(len(doc_xml), idx+140)])
        if '<w:rtl/>' in doc_xml or '<w:rtl' in doc_xml:
            found_rtl = True
            idx = doc_xml.find('<w:rtl')
            print("Found <w:rtl/>. XML Snippet:")
            safe_print(doc_xml[max(0, idx-60):min(len(doc_xml), idx+140)])
            
    if found_bidi and found_rtl:
        print("PASS: w:bidi and w:rtl are present on paragraphs/runs")
    else:
        print(f"FAIL: bidi={found_bidi}, rtl={found_rtl}")

def test_workbook():
    input_json = "scratch/payload_workbook.json"
    output_file = "scratch/test_output.xlsx"
    if not run_export("workbook", input_json, output_file):
        print("FAIL: Workbook export crashed")
        return False
        
    wb = load_workbook(output_file)
    ws = wb.active
    print("\n--- XLSX Verification ---")
    print("sheet_view.rightToLeft =", ws.sheet_view.rightToLeft)
    cell = ws.cell(row=5, column=1)
    safe_print(f"Cell A5 value = '{cell.value}', alignment.horizontal = '{cell.alignment.horizontal}'")
    
    if ws.sheet_view.rightToLeft == True and cell.alignment.horizontal == 'right':
        print("PASS: Sheet rightToLeft is True, cell is right-aligned")
    else:
        print("FAIL: rightToLeft or alignment incorrect")

if __name__ == "__main__":
    test_slides()
    test_document()
    test_workbook()
