import sys
import os
import subprocess
from openpyxl import load_workbook

# Force stdout to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

input_json = "scratch/rtl_test_workbook.json"
output_xlsx = "scratch/rtl_test_workbook.xlsx"

print("Step 1: Running export.py script...")
cmd = [
    "python",
    "packages/api/src/artifacts/export.py",
    "--input", input_json,
    "--output", output_xlsx,
    "--format", "office"
]

res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
if res.returncode != 0:
    print("Export failed!")
    print("STDOUT:", res.stdout)
    print("STDERR:", res.stderr)
    sys.exit(1)

print("Export completed successfully.")

print("Step 2: Loading generated workbook to verify columns...")
wb = load_workbook(output_xlsx)
ws = wb.active

# Let's inspect cell values for Columns A, B, C, D, E
columns = ["A", "B", "C", "D", "E"]
print("\n--- Column Mapping Verification Output ---")
for col in columns:
    # In the generated sheets, headers might start from row 4 or 5 depending on the template KPI/summary row count.
    # Let's find which row has the headers by searching.
    # Let's inspect rows 1 to 10 for this column.
    print(f"Column {col}:")
    for r in range(1, 15):
        val = ws[f"{col}{r}"].value
        if val is not None:
            print(f"  Row {r}: '{val}'")
            
print("\n--- Verification completed! Check the row output mapping above to ensure no off-by-one errors are present. ---")
